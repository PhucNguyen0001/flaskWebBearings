from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from database import *
from base64forfiles import file_to_base64
import io
import openpyxl
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Add this line for flash messages

@app.route('/test')
def test():
    return render_template('themchungloai.html')

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/qlchungloai')
def qlchungloai():
    return render_template('qlchungloai.html', category=Category.select())


@app.route('/qlhangsx')
def qlhangsx():
    return render_template('qlhangsx.html', manufacturer=Manufacturer.select())


@app.route('/qlvongbi')
def qlvongbi():
    return render_template('qlvongbi.html', bearings=select_vongbi())


@app.route('/themchungloai', methods=['GET', "POST"])
def themchungloai():
    if request.method == 'POST':
        categoryName = request.form.get('categoryName')
        file = request.files.get('pictureFile')
        if file:
            try:
                file_content = file.read()
                image_str = f"data:image/{file.filename.split('.')[-1]};base64,{file_to_base64(file_content)}"
                add_category(categoryName, image_str)
                flash("Thêm chủng loại thành công", "success")
                return redirect(url_for('qlchungloai'))
            except Exception as e:
                flash(f"Lỗi: {e}", "error")
                return redirect(url_for('themchungloai'))
        else:
            flash("No file uploaded", "error")
            return redirect(url_for('themchungloai'))
    return render_template('themchungloai.html')


@app.route('/themhangsx', methods=['GET', "POST"])
def themhangsx():
    if request.method == 'POST':
        manufacturerName = request.form.get('manufacturerName')
        file = request.files.get('pictureFile')
        if file:
            try:
                file_content = file.read()
                image_str = f"data:image/{file.filename.split('.')[-1]};base64,{file_to_base64(file_content)}"
                add_manufacturer(manufacturerName, image_str)
                flash("Thêm hãng sản xuất thành công", "success")
                return redirect(url_for('qlhangsx'))
            except Exception as e:
                flash(f"Lỗi: {e}", "error")
                return redirect(url_for('themhangsx'))
        else:
            flash("No file uploaded", "error")
            return redirect(url_for('themhangsx'))
    return render_template('themhangsx.html')


@app.route('/themvongbi', methods=['GET', 'POST'])
def themvongbi():
    if request.method == 'POST':
        try:
            # Get basic bearing info
            product_code = request.form.get('productCode')
            manufacturer_id = request.form.get('manufacturer')
            category_id = request.form.get('category')
            bearing_describe = request.form.get('bearingDescribe')

            # Handle image upload
            file = request.files.get('pictureFile')
            image_str = None
            if file:
                file_content = file.read()
                image_str = f"data:image/{file.filename.split('.')[-1]};base64,{file_to_base64(file_content)}"

            # Collect parameters
            parameters = {}
            for key, value in request.form.items():
                if key.startswith('parameter-') and value:
                    parameter_id = int(key.split('-')[1])
                    parameters[parameter_id] = float(value)

            # Add bearing to database
            new_bearing = add_bearing(product_code, manufacturer_id, category_id, bearing_describe, image_str, parameters)

            flash("Thêm vòng bi thành công", "success")
            return redirect(url_for('qlvongbi'))
        except Exception as e:
            flash(f"Lỗi: {e}", "error")
            return redirect(url_for('themvongbi'))

    return render_template('themvongbi.html', 
                           listCategory=Category.select(Category.idCategory, Category.categoryName),
                           listManufacturer=Manufacturer.select(Manufacturer.idManufacturer, Manufacturer.manufacturerName),
                           listParameter=Parameter.select())


@app.route('/suavongbi/<int:bearing_id>', methods=['GET', 'POST'])
def suavongbi(bearing_id):
    bearing = get_bearing_by_id(bearing_id)
    if request.method == 'POST':
        try:
            # Get updated bearing info
            product_code = request.form.get('productCode')
            manufacturer_id = request.form.get('manufacturer')
            category_id = request.form.get('category')
            bearing_describe = request.form.get('description')

            # Handle image upload
            file = request.files.get('pictureFile')
            image_str = None
            if file:
                file_content = file.read()
                image_str = f"data:image/{file.filename.split('.')[-1]};base64,{file_to_base64(file_content)}"

            # Collect parameters
            parameters = {}
            for key, value in request.form.items():
                if key.startswith('parameter-') and value:
                    parameter_id = int(key.split('-')[1])
                    parameters[parameter_id] = float(value)

            # Update bearing in database
            update_bearing(bearing_id, product_code, manufacturer_id, category_id, bearing_describe, image_str, parameters)

            flash("Cập nhật vòng bi thành công", "success")
            return redirect(url_for('qlvongbi'))
        except Exception as e:
            flash(f"Lỗi: {e}", "error")
            return redirect(url_for('suavongbi', bearing_id=bearing_id))

    # Create parameter_dict for GET request
    parameter_dict = {bp.idParameter.idParameter: bp.value for bp in bearing.parameters}

    return render_template('suavongbi.html', 
                           bearing=bearing,
                           listCategory=Category.select(Category.idCategory, Category.categoryName),
                           listManufacturer=Manufacturer.select(Manufacturer.idManufacturer, Manufacturer.manufacturerName),
                           listParameter=Parameter.select(),
                           parameter_dict=parameter_dict)


@app.route('/suachungloai/<int:category_id>', methods=['GET', 'POST'])
def suachungloai(category_id):
    category = Category.get_by_id(category_id)
    if request.method == 'POST':
        categoryName = request.form.get('categoryName')
        file = request.files.get('pictureFile')
        try:
            if file:
                file_content = file.read()
                image_str = f"data:image/{file.filename.split('.')[-1]};base64,{file_to_base64(file_content)}"
                category.categoryPicture = image_str
            category.categoryName = categoryName
            category.save()
            flash("Cập nhật chủng loại thành công", "success")
            return redirect(url_for('qlchungloai'))
        except Exception as e:
            flash(f"Lỗi: {e}", "error")
            return redirect(url_for('suachungloai', category_id=category_id))
    return render_template('suachungloai.html', category=category)


@app.route('/suahangsx/<int:manufacturer_id>', methods=['GET', 'POST'])
def suahangsx(manufacturer_id):
    manufacturer = Manufacturer.get_by_id(manufacturer_id)
    if request.method == 'POST':
        manufacturerName = request.form.get('manufacturerName')
        file = request.files.get('pictureFile')
        try:
            if file:
                file_content = file.read()
                image_str = f"data:image/{file.filename.split('.')[-1]};base64,{file_to_base64(file_content)}"
                manufacturer.manufacturerPicture = image_str
            manufacturer.manufacturerName = manufacturerName
            manufacturer.save()
            flash("Cập nhật hãng sản xuất thành công", "success")
            return redirect(url_for('qlhangsx'))
        except Exception as e:
            flash(f"Lỗi: {e}", "error")
            return redirect(url_for('suahangsx', manufacturer_id=manufacturer_id))
    return render_template('suahangsx.html', manufacturer=manufacturer)


@app.route('/upload_bearings', methods=['GET', 'POST'])
def upload_bearings():
    if request.method == 'GET':
        # Lấy danh sách parameters, manufacturers và categories để hiển thị trên trang
        parameters = Parameter.select()
        manufacturers = Manufacturer.select()
        categories = Category.select()
        
        return render_template('upload_bearings.html', 
                               parameters=parameters, 
                               manufacturers=manufacturers, 
                               categories=categories)
    
    elif request.method == 'POST':
        try:
            file = request.files['excelFile']  # Change 'file' to 'excelFile'
            if file and file.filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active

                headers = [cell.value for cell in ws[1]]
                parameter_columns = {col: headers.index(col) for col in headers if ' - ' in col}

                for row in ws.iter_rows(min_row=2, values_only=True):
                    product_code = row[headers.index('mã vòng')]
                    manufacturer_id = int(row[headers.index('nhà sx')])  # Change to ID
                    category_id = int(row[headers.index('chủng loại')])  # Change to ID
                    description = row[headers.index('mô tả')]

                    # Collect parameters
                    parameters = {}
                    for col, index in parameter_columns.items():
                        parameter_id = int(col.split(' - ')[0])
                        value = row[index]
                        if value is not None:
                            parameters[parameter_id] = float(value)

                    # Add bearing with collected parameters
                    bearing = add_bearing(product_code, manufacturer_id, category_id, description, None, parameters)

                    # Add parameters
                    for col, index in parameter_columns.items():
                        parameter_id = int(col.split(' - ')[0])
                        value = row[index]
                        if value is not None:
                            BearingParameter.create(idBearing=bearing, idParameter=parameter_id, value=float(value))

                flash("Tải lên và thêm vòng bi thành công", "success")
            else:
                flash("Vui lòng tải lên file Excel (.xlsx)", "error")
        except Exception as e:
            flash(f"Lỗi khi xử lý file: {str(e)}", "error")

        return redirect(url_for('upload_bearings'))

    # ... (rest of the function remains the same) ...

@app.route('/download_template')
def download_template():
    # Tạo một workbook mới
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Thêm tiêu đề cột
    headers = ['mã vòng', 'nhà sx', 'chủng loại', 'mô tả']
    parameters = Parameter.select()
    for param in parameters:
        headers.append(f"{param.idParameter} - {param.describe}")
    
    ws.append(headers)
    
    # Lưu workbook vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, 
                     download_name='bearing_template.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    app.run(debug=True)